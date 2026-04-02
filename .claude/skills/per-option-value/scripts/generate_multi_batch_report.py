from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Load template
template_path = r"c:\Users\tommy\Documents\Github\simple-agent\.claude\skills\per-option-value\template\ESO multi-batch.xlsx"
wb = load_workbook(template_path)
ws = wb.active

# Insert new column for Batch 3 (after column D)
ws.insert_cols(5)  # Insert column E

# Copy formatting from column D to E
for row in range(1, 35):
    source_cell = ws.cell(row=row, column=4)
    target_cell = ws.cell(row=row, column=5)
    
    # Copy value/formula
    if source_cell.value:
        value = str(source_cell.value)
        # Update column references from D to E
        if value.startswith('='):
            value = value.replace('D', 'E')
        target_cell.value = value
    
    # Copy formatting
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )
    if source_cell.fill:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical
        )
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format

# Update header for Batch 3
ws.cell(row=14, column=5).value = "Batch 3"

# Data mapping
data = {
    # Report Information
    "{{Valuation Subject}}": "aaa",
    "{{Valuation Date}}": "12 Dec 2023",
    "{{Grant Date of the Subject}}": "12 Dec 2023",
    "{{Maturity Date}}": "12 Dec 2024",
    "{{Total No. of Share Options}}": "5,200,000",  # Sum of all batches
    
    # Batch 1
    "{{No. of Share Options1}}": "1,650,000",
    "{{Vesting Date1}}": "12 Dec 2023",
    "{{Spot Price}}": "0.066",
    "{{Strike Price}}": "0.066",
    "{{Volatility}}": "64.23%",
    "{{Risk-free Rate}}": "0%",
    "{{Dividend Yield}}": "0%",
    "{{Exercise Multiple}}": "2.8",
    "{{Post-vesting Exit Rate}}": "0%",
    "{{Time to Vest1}}": "0.0",
    "{{PerOptionValue1}}": "0.017",
    
    # Batch 2
    "{{No. of Share Options2}}": "1,650,000",
    "{{Vesting Date2}}": "12 Jan 2024",
    "{{Time to Vest2}}": "0.085",
    "{{PerOptionValue2}}": "0.017",
    
    # Batch 3
    "{{No. of Share Options3}}": "1,900,000",
    "{{Vesting Date3}}": "11 Dec 2024",
    "{{Time to Vest3}}": "0.997",
    "{{PerOptionValue3}}": "0.017",
}

# Replace placeholders in all cells
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            for placeholder, value in data.items():
                if placeholder in str(cell.value):
                    cell.value = str(cell.value).replace(placeholder, value)

# Update Batch 3 specific cells
ws.cell(row=15, column=5).value = "1,900,000"  # No. of Share Options
ws.cell(row=16, column=5).value = "11 Dec 2024"  # Vesting Date
ws.cell(row=26, column=5).value = "0.997"  # Time to Vest
ws.cell(row=27, column=5).value = "0.017"  # PerOptionValue
ws.cell(row=28, column=5).value = "=E27*E15"  # Total Value formula
ws.cell(row=29, column=5).value = "=E27/E17"  # Ratio formula

# Save report
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_dir = r"c:\Users\tommy\Documents\Github\simple-agent\.claude\skills\per-option-value\reports"
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, f"ESO_Valuation_Report_UniBioScienceGroup_{timestamp}.xlsx")
wb.save(output_path)
print(f"Report saved to: {output_path}")
